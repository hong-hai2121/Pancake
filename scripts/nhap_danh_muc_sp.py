"""Nhập danh mục sản phẩm từ `docs/Bang-san-pham-CRM-hoan-thien.xlsx` vào B6.

Đi qua `product_service` chứ KHÔNG insert thẳng — để mỗi sản phẩm có ngay
phiên bản 1 (chốt giá + nội dung, FR-060 "đổi giá tạo phiên bản mới") và có
vết audit như người dùng thao tác trên web.

Ánh xạ cột Excel -> CRM:
    Mã sản phẩm                    -> products.product_code
    Tên sản phẩm chuẩn             -> products.name
    Giá bán                        -> products.price + product_versions.price
    Quy cách                       -> products.package (+ tách số -> units_per_package)
    Nhóm vấn đề                    -> products.product_type
    Cách dùng                      -> product_versions.usage_text
    Thành phần / Mô tả / Nội dung tư vấn / Vai trò / Tài liệu / Kịch bản giá
    + sheet "Nội dung mở rộng"     -> product_versions.approved_claims (gắn nhãn)
    Trạng thái kiểm duyệt          -> products.approval_status

NỘI DUNG CẤM: file nguồn KHÔNG có cột này -> `prohibited_claims` để rỗng.
Danh sách câu từ cấm nằm ở file `Quy-trinh-cham-soc-ban-lai-chuan-hoa-CRM.xlsx`
(sheet "Quy định"), nhập riêng khi làm FR-143.

Idempotent: chạy lại thì mã đã có sẽ được CẬP NHẬT (giá đổi -> tự sinh phiên
bản mới), không tạo trùng.

Chạy:
    python scripts/nhap_danh_muc_sp.py            # xem trước, KHÔNG ghi
    python scripts/nhap_danh_muc_sp.py --ghi      # ghi thật
"""

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl                                          # noqa: E402

from app.db.repositories import catalog_repo             # noqa: E402
from app.services import product_service                 # noqa: E402

NGUON = Path(__file__).resolve().parents[1] / "docs" / \
    "Bang-san-pham-CRM-hoan-thien.xlsx"
SHEET_SP = "Danh mục sản phẩm CRM"
SHEET_MO_RONG = "Nội dung mở rộng"

# Sheet "Danh mục cấu hình" của chính file nguồn quy định 4 trạng thái này.
TRANG_THAI_DUYET = {
    "chờ duyệt chuyên môn": ("pending", "active"),
    "đã duyệt": ("approved", "active"),
    "tạm khóa": ("draft", "inactive"),
    "ngừng sử dụng": ("draft", "discontinued"),
}

# Cột nội dung -> nhãn ghi kèm trong approved_claims (giữ NGUYÊN VĂN, không tự
# diễn giải y khoa — đúng nguyên tắc ghi trong sheet "Danh mục cấu hình").
# Sheet "Danh mục cấu hình" cột "Nhóm vấn đề gợi ý". Giá trị lạ vẫn NHẬP
# (không tự bịa/bỏ dữ liệu nguồn) nhưng có cảnh báo để người phụ trách sửa lại.
NHOM_CHUAN = {"trào", "viêm", "trào + viêm", "đại tràng", "khác"}

COT_NOI_DUNG = [
    ("Thành phần - công dụng", "Thành phần – công dụng"),
    ("Mô tả", "Mô tả"),
    ("Nội dung tư vấn khách hàng", "Nội dung tư vấn khách hàng"),
    ("Mô tả vai trò trong liệu trình", "Vai trò trong liệu trình"),
    ("Kịch bản giá / phác đồ", "Kịch bản giá / phác đồ"),
    ("Tài liệu liên quan / công bố", "Tài liệu công bố"),
]


def _chuoi(v) -> str:
    return "" if v is None else str(v).replace("\xa0", " ").strip()


def _so_tien(v) -> float | None:
    s = re.sub(r"[^\d.]", "", _chuoi(v).replace(",", ""))
    try:
        return float(s) if s else None
    except ValueError:
        return None


def _don_vi(quy_cach: str) -> int | None:
    """'60 viên / hộp' -> 60. Không rõ thì None (cột này check > 0)."""
    m = re.search(r"\d+", quy_cach)
    if not m:
        return None
    n = int(m.group())
    return n if n > 0 else None


def _doc_excel() -> list[dict]:
    wb = openpyxl.load_workbook(NGUON, data_only=True)
    ws = wb[SHEET_SP]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_chuoi(h) for h in rows[0]]

    mo_rong: dict[str, str] = {}
    if SHEET_MO_RONG in wb.sheetnames:
        wr = wb[SHEET_MO_RONG]
        rr = list(wr.iter_rows(values_only=True))
        h2 = [_chuoi(h) for h in rr[0]]
        for r in rr[1:]:
            d = dict(zip(h2, r))
            ma = _chuoi(d.get("Mã sản phẩm"))
            noi_dung = _chuoi(d.get("Nội dung chuyên sâu / bổ sung"))
            if ma and noi_dung:
                mo_rong[ma] = noi_dung

    ra = []
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        ma = _chuoi(d.get("Mã sản phẩm"))
        ten = _chuoi(d.get("Tên sản phẩm chuẩn"))
        if not ma or not ten:
            continue
        quy_cach = _chuoi(d.get("Quy cách"))
        claims = []
        for cot, nhan in COT_NOI_DUNG:
            noi_dung = _chuoi(d.get(cot))
            if noi_dung:
                claims.append(f"{nhan}: {noi_dung}")
        if mo_rong.get(ma):
            claims.append(f"Nội dung chuyên sâu / bổ sung: {mo_rong[ma]}")

        duyet, ban = TRANG_THAI_DUYET.get(
            _chuoi(d.get("Trạng thái kiểm duyệt")).lower(), ("pending", "active"))
        ra.append({
            "product_code": ma,
            "name": ten,
            "price": _so_tien(d.get("Giá bán")),
            "package": quy_cach or None,
            "units_per_package": _don_vi(quy_cach),
            "product_type": _chuoi(d.get("Nhóm vấn đề")) or None,
            "usage_text": _chuoi(d.get("Cách dùng")) or None,
            "approved_claims": claims,
            "prohibited_claims": [],
            "_approval": duyet,
            "_status": ban,
        })
    return ra


def _theo_ma() -> dict[str, dict]:
    da_co, _ = catalog_repo.list_products(limit=1000, offset=0)
    return {sp["product_code"]: sp for sp in da_co if sp.get("product_code")}


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--ghi", action="store_true",
                   help="Ghi thật vào DB (mặc định chỉ xem trước)")
    args = p.parse_args()

    sps = _doc_excel()
    print(f"Đọc {NGUON.name}: {len(sps)} sản phẩm\n")

    la = [(sp["product_code"], sp["product_type"]) for sp in sps
          if sp["product_type"] and sp["product_type"].lower() not in NHOM_CHUAN]
    thieu = [sp["product_code"] for sp in sps if not sp["product_type"]]

    if not args.ghi:
        for sp in sps:
            gia = f"{sp['price']:,.0f}đ" if sp["price"] else "—"
            print(f"  {sp['product_code']}  {sp['name'][:32]:34s} {gia:>12} "
                  f"| {sp['package'] or '—':16s} | đv/hộp={sp['units_per_package']} "
                  f"| nhóm={sp['product_type'] or '—'} "
                  f"| {len(sp['approved_claims'])} khối nội dung "
                  f"| duyệt={sp['_approval']}")
        print("\nXem trước — CHƯA ghi gì. Chạy lại với --ghi để nhập thật.")
        return 0

    da_co = _theo_ma()
    actor = None      # script chạy tay: audit ghi user_id = null (không mạo danh ai)
    tao, cap_nhat, loi = 0, 0, 0
    for sp in sps:
        ma = sp["product_code"]
        truong = {k: v for k, v in sp.items() if not k.startswith("_")}
        try:
            if ma in da_co:
                pid = da_co[ma]["id"]
                product_service.update_product(
                    pid, {k: truong[k] for k in
                          ("name", "price", "package", "units_per_package",
                           "product_type")}, actor=actor)
                product_service.add_version(
                    pid, {"price": truong["price"],
                          "usage_text": truong["usage_text"],
                          "approved_claims": truong["approved_claims"],
                          "prohibited_claims": truong["prohibited_claims"]},
                    actor=actor)
                cap_nhat += 1
                print(f"  ~ {ma} cập nhật + phiên bản nội dung mới")
            else:
                moi = product_service.create_product(truong, actor=actor)
                pid = moi["id"]
                tao += 1
                print(f"  + {ma} {moi['name'][:40]}")
            # Trạng thái duyệt/bán theo đúng cột Excel (create mặc định 'draft')
            catalog_repo.set_product_approval(pid, sp["_approval"])
            if sp["_status"] != "active":
                catalog_repo.set_product_status(pid, sp["_status"])
        except Exception as err:            # noqa: BLE001 — 1 dòng hỏng không chặn cả mẻ
            loi += 1
            print(f"  ! {ma} LỖI: {type(err).__name__}: {err}")

    print(f"\nXong: tạo mới {tao} · cập nhật {cap_nhat} · lỗi {loi}")
    _canh_bao(la, thieu)
    return 1 if loi else 0


def _canh_bao(la: list[tuple[str, str]], thieu: list[str]) -> None:
    """Nhắc phần dữ liệu nguồn cần người phụ trách sửa lại trên màn 42."""
    if la:
        print("\n⚠ Nhóm vấn đề KHÔNG thuộc danh mục chuẩn (đã nhập nguyên văn, "
              "sửa lại ở màn 42 /crm/san-pham):")
        for ma, v in la:
            print(f"    {ma}: {v!r}")
    if thieu:
        print(f"\n⚠ {len(thieu)} sản phẩm THIẾU nhóm vấn đề "
              f"({', '.join(thieu)}) — rule engine B6 lọc theo nhóm này, "
              "thiếu thì mẫu liệu trình khó khớp đúng.")


if __name__ == "__main__":
    sys.exit(main())
